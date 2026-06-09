from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from industries.pizza_calc import calculate, get_default_input
from excel_validation import pizza_default_validation_report, pizza_excel_updates_from_model, recalc_excel_output


TOLERANCE = 1e-6
REPORT_PATH = Path("/workspace/output/pizza_exact_validation_report.xlsx")


@dataclass(frozen=True)
class RegressionResult:
    name: str
    description: str
    python_value: float
    excel_value: float

    @property
    def diff(self) -> float:
        return self.python_value - self.excel_value

    @property
    def passed(self) -> bool:
        return abs(self.diff) <= TOLERANCE


def regression_cases() -> list[tuple[str, str, Callable]]:
    return [
        ("default", "원본 기본 입력값", lambda data: None),
        ("weekday_friday", "조사요일 월 -> 금", lambda data: setattr(data, "weekday", "금")),
        ("month_3", "조사월 7 -> 3", lambda data: setattr(data, "survey_month", 3)),
        ("traffic_d40_f40_up", "12~13시 20대 남 통행량 0 -> 80", lambda data: data.traffic[1].__setitem__(2, 80)),
        (
            "households_4500",
            "아파트 3,200 / 주택계 4,500",
            lambda data: (setattr(data, "apartment_households", 3200), setattr(data, "total_households", 4500)),
        ),
        ("worker_5000", "상주인구 910 -> 5,000", lambda data: setattr(data, "worker_population", 5000)),
        (
            "apt_size_low",
            "30평 이상 아파트 비율 하락",
            lambda data: data.extra_inputs["apartment_size_mix"].__setitem__(
                0,
                {"0~20평": 300, "20~29평": 300, "30~39평": 30, "40~49평": 20, "50평 이상": 0, "단독/다세대": 0},
            ),
        ),
        (
            "apt_price_low",
            "3억원 이상 아파트 가격대 비율 하락",
            lambda data: data.extra_inputs["apartment_price_mix"].__setitem__(
                0,
                {"1억 미만": 300, "1억원대": 200, "2억원대": 200, "3억원대": 10, "4억원대": 10, "5억원대": 0, "6억원 이상": 0},
            ),
        ),
        ("candidate_no_delivery", "후보점 배달 여부 1 -> 0", lambda data: data.extra_inputs["direct_competitor_rows"][0].update({"배달 여부": 0})),
        ("candidate_no_takeout", "후보점 테이크아웃 여부 1 -> 0", lambda data: data.extra_inputs["direct_competitor_rows"][0].update({"테이크아웃 여부": 0})),
        ("candidate_area_10", "후보점 면적 36평 -> 10평", lambda data: setattr(data.direct_competitors[0], "area", 10)),
        ("competitor_distance_300", "경쟁점 거리 158m -> 300m", lambda data: setattr(data.direct_competitors[1], "distance", 300)),
    ]


def run_regressions(verbose: bool = False) -> list[RegressionResult]:
    results: list[RegressionResult] = []
    for name, description, mutator in regression_cases():
        if verbose:
            print(f"Running {name}: {description}", flush=True)
        data = get_default_input()
        mutator(data)
        python_value = calculate(data)["daily_sales_thousand"]
        excel_value = recalc_excel_output("pizza", pizza_excel_updates_from_model(data), timeout_seconds=30)
        result = RegressionResult(name, description, python_value, excel_value)
        results.append(result)
        if verbose:
            print({"name": result.name, "diff": result.diff, "passed": result.passed}, flush=True)
    return results


def run_default_intermediate_checks() -> list[dict]:
    result = calculate(get_default_input())
    return pizza_default_validation_report(result, tolerance=TOLERANCE)


def build_report(intermediate: list[dict], regressions: list[RegressionResult]) -> Path:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    details = workbook.create_sheet("Intermediate Checks")
    regression_sheet = workbook.create_sheet("Regression Cases")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")

    summary.append(["항목", "결과"])
    summary.append(["업종", "피자"])
    summary.append(["기본 중간 수식 통과", f"{sum(row['passed'] for row in intermediate)} / {len(intermediate)}"])
    summary.append(["회귀 테스트 통과", f"{sum(row.passed for row in regressions)} / {len(regressions)}"])
    summary.append(["허용 오차", TOLERANCE])
    summary.append(["원본 엑셀 최종 셀", "표지!F12"])

    details.append(["name", "python_value", "excel_value", "diff", "passed"])
    for row in intermediate:
        details.append([row["name"], row["python_value"], row["excel_value"], row["diff"], row["passed"]])

    regression_sheet.append(["name", "description", "python_value", "excel_value", "diff", "passed"])
    for row in regressions:
        regression_sheet.append([row.name, row.description, row.python_value, row.excel_value, row.diff, row.passed])

    for sheet in [summary, details, regression_sheet]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in sheet.iter_rows(min_row=2):
            passed_cell = row[-1]
            if passed_cell.value is True:
                passed_cell.fill = pass_fill
            elif passed_cell.value is False:
                passed_cell.fill = fail_fill
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 12), 60)
        sheet.freeze_panes = "A2"

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(REPORT_PATH)
    return REPORT_PATH


def main() -> None:
    intermediate = run_default_intermediate_checks()
    regressions = run_regressions(verbose=True)
    print("Intermediate checks")
    print(f"{sum(row['passed'] for row in intermediate)} / {len(intermediate)} passed")
    for row in intermediate:
        if not row["passed"]:
            print(row)
    print("\nRegression checks")
    print(f"{sum(row.passed for row in regressions)} / {len(regressions)} passed")
    for row in regressions:
        print(
            {
                "name": row.name,
                "description": row.description,
                "python_value": row.python_value,
                "excel_value": row.excel_value,
                "diff": row.diff,
                "passed": row.passed,
            }
        )
    print(f"\nReport: {build_report(intermediate, regressions)}")


if __name__ == "__main__":
    main()
