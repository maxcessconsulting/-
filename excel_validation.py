from __future__ import annotations

import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from industries.haejang_excel_map import default_parity_checks, model_input_to_excel_updates
    from industries.chicken_excel_map import (
        default_parity_checks as chicken_default_parity_checks,
        model_input_to_excel_updates as chicken_model_input_to_excel_updates,
    )
    from industries.pizza_excel_map import (
        default_parity_checks as pizza_default_parity_checks,
        model_input_to_excel_updates as pizza_model_input_to_excel_updates,
    )
except ModuleNotFoundError:
    from haejang_excel_map import default_parity_checks, model_input_to_excel_updates
    from chicken_excel_map import (
        default_parity_checks as chicken_default_parity_checks,
        model_input_to_excel_updates as chicken_model_input_to_excel_updates,
    )
    from pizza_excel_map import (
        default_parity_checks as pizza_default_parity_checks,
        model_input_to_excel_updates as pizza_model_input_to_excel_updates,
    )


@dataclass(frozen=True)
class ExcelTarget:
    workbook_path: str
    output_sheet: str
    output_cell: str
    input_sheet: str = "조사치 입력 장표"


TARGETS = {
    "haejang": ExcelTarget("/workspace/user_files/01-pgm.xlsx", "표지", "D9"),
    "pizza": ExcelTarget("/workspace/user_files/01-pizza.xlsx", "표지", "F12"),
    "chicken": ExcelTarget("/workspace/user_files/01-chicken.xlsx", "표지", "F11"),
}


def recalc_excel_output(industry_code: str, cell_updates: dict[str, Any] | None = None, timeout_seconds: int = 30) -> float:
    target = TARGETS[industry_code]
    source = Path(target.workbook_path)
    if not source.exists():
        raise FileNotFoundError(f"원본 엑셀 파일을 찾을 수 없습니다: {source}")

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / source.name
        shutil.copy2(source, temp_path)

        if cell_updates:
            workbook = load_workbook(temp_path)
            sheet = workbook[target.input_sheet]
            for cell, value in cell_updates.items():
                sheet[cell] = value
            workbook.save(temp_path)

        output_dir = Path(temp_dir) / "out"
        output_dir.mkdir()
        user_installation = Path(temp_dir) / "lo"
        subprocess.run(
            [
                "soffice",
                "--headless",
                f"-env:UserInstallation=file://{user_installation}",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(output_dir),
                str(temp_path),
            ],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout_seconds,
        )
        recalculated = output_dir / temp_path.name
        if not recalculated.exists():
            recalculated = temp_path

        workbook = load_workbook(recalculated, data_only=True)
        value = workbook[target.output_sheet][target.output_cell].value
        if value is None:
            raise ValueError(f"{target.output_sheet}!{target.output_cell} 값이 비어 있습니다.")
        return float(value)


def assert_close_to_excel(industry_code: str, python_value: float, cell_updates: dict[str, Any] | None = None, tolerance: float = 1e-6) -> None:
    excel_value = recalc_excel_output(industry_code, cell_updates)
    diff = abs(excel_value - python_value)
    if diff > tolerance:
        raise AssertionError(
            f"{industry_code} 불일치: excel={excel_value}, python={python_value}, diff={diff}, tolerance={tolerance}"
        )


def haejang_default_validation_report(tolerance: float = 1e-6) -> list[dict[str, float | str | bool]]:
    report = []
    for check in default_parity_checks():
        report.append(
            {
                "name": check.name,
                "python_value": check.python_value,
                "excel_value": check.excel_value,
                "diff": check.diff,
                "passed": abs(check.diff) <= tolerance,
            }
        )
    return report


def haejang_excel_updates_from_model(data) -> dict[str, Any]:
    return model_input_to_excel_updates(data)


def chicken_default_validation_report(result: dict, tolerance: float = 1e-6) -> list[dict[str, float | str | bool]]:
    report = []
    for check in chicken_default_parity_checks(result):
        report.append(
            {
                "name": check.name,
                "python_value": check.python_value,
                "excel_value": check.excel_value,
                "diff": check.diff,
                "passed": abs(check.diff) <= tolerance,
            }
        )
    return report


def chicken_excel_updates_from_model(data) -> dict[str, Any]:
    return chicken_model_input_to_excel_updates(data)


def pizza_default_validation_report(result: dict, tolerance: float = 1e-6) -> list[dict[str, float | str | bool]]:
    report = []
    for check in pizza_default_parity_checks(result):
        report.append(
            {
                "name": check.name,
                "python_value": check.python_value,
                "excel_value": check.excel_value,
                "diff": check.diff,
                "passed": abs(check.diff) <= tolerance,
            }
        )
    return report


def pizza_excel_updates_from_model(data) -> dict[str, Any]:
    return pizza_model_input_to_excel_updates(data)
