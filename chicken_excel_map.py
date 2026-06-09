from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pure_model import AGE_COLUMNS, ModelInput


WORKBOOK_PATH = Path("/workspace/user_files/01-chicken.xlsx")
INPUT_SHEET = "조사치 입력 장표"
OUTPUT_SHEET = "표지"
OUTPUT_DAILY_SALES_CELL = "F11"

REGION_TO_EXCEL_CELL = {
    "서울 경기": "B11",
    "충청권": "D11",
    "호남권": "F11",
    "대구경북": "H11",
    "부산 경남": "J11",
    "강원권": "L11",
}

APARTMENT_SIZE_TO_CELL = {
    "0~20평": "D28",
    "20~29평": "E28",
    "30~39평": "F28",
    "40~49평": "G28",
    "50평 이상": "H28",
}

APARTMENT_PRICE_TO_CELL = {
    "1억 미만": "D32",
    "1억원대": "E32",
    "2억원대": "F32",
    "3억원대": "G32",
    "4억원대": "H32",
    "5억원대": "I32",
    "6억원 이상": "J32",
}

TRAFFIC_ROW_BY_MODEL_INDEX = {
    4: 39,   # 15~16
    5: 40,   # 16~17
    6: 41,   # 17~18
    7: 42,   # 18~19
    8: 43,   # 19~20
    9: 44,   # 20~21
    10: 45,  # 21~22
}

DIRECT_COMPETITOR_START_ROW = 56
INDIRECT_COMPETITOR_START_ROW = 67
COMPETITOR_COLUMNS = {
    "name": "D",
    "area": "E",
    "distance": "F",
    "location": "G",
    "visibility": "H",
    "accessibility": "I",
    "floor": "J",
    "sides": "K",
    "frontage": "L",
    "facility": "M",
    "parking": "N",
    "price": "O",
}

SERVICE_COLUMNS = {
    "홀판매여부": "P",
    "배달 여부": "Q",
    "테이크아웃 여부": "R",
}


@dataclass(frozen=True)
class ExcelInputCell:
    field: str
    sheet: str
    cell: str
    label: str
    required: bool = True


@dataclass(frozen=True)
class ExcelParityCheck:
    name: str
    python_value: float
    excel_value: float

    @property
    def diff(self) -> float:
        return self.python_value - self.excel_value


BASE_INPUT_CELLS = [
    ExcelInputCell("survey_month", INPUT_SHEET, "D6", "조사월"),
    ExcelInputCell("weekday", INPUT_SHEET, "K6", "조사요일"),
    ExcelInputCell("region", INPUT_SHEET, "J11", "지역 권역"),
    ExcelInputCell("apartment_households", INPUT_SHEET, "D22", "아파트 세대수"),
    ExcelInputCell("single_households", INPUT_SHEET, "F22", "단독/다세대 세대수"),
    ExcelInputCell("resident_population", INPUT_SHEET, "J22", "주거인구"),
    ExcelInputCell("worker_population", INPUT_SHEET, "L22", "상주인구"),
    ExcelInputCell("deposit", INPUT_SHEET, "C83", "임차보증금"),
    ExcelInputCell("goodwill", INPUT_SHEET, "D83", "영업권"),
    ExcelInputCell("monthly_rent", INPUT_SHEET, "E83", "월임대료"),
]


def base_input_cell_map() -> dict[str, str]:
    mapping = {item.field: f"{item.sheet}!{item.cell}" for item in BASE_INPUT_CELLS}
    mapping.update({f"apartment_size_mix.{key}": f"{INPUT_SHEET}!{cell}" for key, cell in APARTMENT_SIZE_TO_CELL.items()})
    mapping.update({f"apartment_price_mix.{key}": f"{INPUT_SHEET}!{cell}" for key, cell in APARTMENT_PRICE_TO_CELL.items()})
    mapping["direct_competitors"] = f"{INPUT_SHEET}!D56:R65"
    mapping["indirect_competitors"] = f"{INPUT_SHEET}!D67:R76"
    mapping["traffic"] = f"{INPUT_SHEET}!D39:O45"
    return mapping


def model_input_to_excel_updates(data: ModelInput) -> dict[str, Any]:
    updates: dict[str, Any] = {
        "D6": data.survey_month,
        "K6": data.weekday,
        "D22": data.apartment_households,
        "F22": max(data.total_households - data.apartment_households, 0),
        "J22": data.resident_population,
        "L22": data.worker_population,
        "C83": data.deposit,
        "D83": data.goodwill,
        "E83": data.monthly_rent,
    }
    updates.update(_region_updates(data.region))
    updates.update(_market_mix_updates(data, "apartment_size_mix", APARTMENT_SIZE_TO_CELL))
    updates.update(_market_mix_updates(data, "apartment_price_mix", APARTMENT_PRICE_TO_CELL))
    updates.update(_traffic_updates(data))
    updates.update(_competitor_updates(data.direct_competitors, DIRECT_COMPETITOR_START_ROW, _extra_rows(data, "direct_competitor_rows")))
    updates.update(_competitor_updates(data.indirect_competitors, INDIRECT_COMPETITOR_START_ROW, _extra_rows(data, "indirect_competitor_rows")))
    return updates


def _region_updates(region: str) -> dict[str, int]:
    updates = {cell: 0 for cell in REGION_TO_EXCEL_CELL.values()}
    selected_cell = REGION_TO_EXCEL_CELL.get(region)
    if selected_cell:
        updates[selected_cell] = 1
    return updates


def _market_mix_updates(data: ModelInput, key: str, cell_map: dict[str, str]) -> dict[str, Any]:
    rows = _extra_rows(data, key)
    if not rows:
        return {}
    first_row = rows[0]
    return {cell: first_row.get(column) for column, cell in cell_map.items() if first_row.get(column) is not None}


def _traffic_updates(data: ModelInput) -> dict[str, Any]:
    updates: dict[str, Any] = {}
    for model_index, excel_row in TRAFFIC_ROW_BY_MODEL_INDEX.items():
        if model_index >= len(data.traffic):
            continue
        for offset, value in enumerate(data.traffic[model_index][: len(AGE_COLUMNS)]):
            updates[f"{chr(68 + offset)}{excel_row}"] = value
    return updates


def _competitor_updates(competitors, start_row: int, extra_rows: list[dict[str, Any]]) -> dict[str, Any]:
    updates: dict[str, Any] = {}
    for offset, competitor in enumerate(competitors[:10]):
        row = start_row + offset
        for attr, column in COMPETITOR_COLUMNS.items():
            if start_row == DIRECT_COMPETITOR_START_ROW and offset == 0 and attr == "distance":
                continue
            updates[f"{column}{row}"] = getattr(competitor, attr)
        extra_row = extra_rows[offset] if offset < len(extra_rows) else {}
        for service_name, column in SERVICE_COLUMNS.items():
            value = extra_row.get(service_name)
            if value is not None:
                updates[f"{column}{row}"] = value
    return updates


def _extra_rows(data: ModelInput, key: str) -> list[dict[str, Any]]:
    extra_inputs = getattr(data, "extra_inputs", {}) or {}
    rows = extra_inputs.get(key, [])
    if hasattr(rows, "to_dict"):
        rows = rows.to_dict("records")
    return [dict(row) for row in rows]


def read_stored_daily_sales(workbook_path: Path = WORKBOOK_PATH) -> float:
    workbook = load_workbook(workbook_path, data_only=True)
    value = workbook[OUTPUT_SHEET][OUTPUT_DAILY_SALES_CELL].value
    if value is None:
        raise ValueError(f"{OUTPUT_SHEET}!{OUTPUT_DAILY_SALES_CELL} 값이 비어 있습니다.")
    return float(value)


def default_parity_checks(result: dict, workbook_path: Path = WORKBOOK_PATH) -> list[ExcelParityCheck]:
    workbook = load_workbook(workbook_path, data_only=True)
    cells = [
        ("daily_sales_thousand", result["daily_sales_thousand"], "표지", "F11"),
        ("daily_traffic", result["daily_traffic"], "기초 데이터(입력불가)", "C14"),
        ("main_customer_ratio", result["main_customer_ratio"], "조사치 입력 장표", "M16"),
        ("takeout_potential", result["traffic_potential_total"], "SIMULATION(입력불가)", "B63"),
        ("delivery_potential", result["household_potential_total"], "SIMULATION(입력불가)", "C63"),
        ("hall_potential", result["worker_potential_total"], "SIMULATION(입력불가)", "D63"),
        ("candidate_takeout_sales", result["candidate_traffic_sales"], "SIMULATION(입력불가)", "G67"),
        ("candidate_delivery_sales", result["candidate_household_sales"], "SIMULATION(입력불가)", "H67"),
        ("candidate_hall_sales", result["candidate_worker_sales"], "SIMULATION(입력불가)", "I67"),
        ("pre_date_sales", result["channel_potentials"]["pre_date_sales"], "SIMULATION(입력불가)", "H62"),
        ("month_index", result["month_index"], "기초자료", "B208"),
        ("weekday_index", result["weekday_index"], "기초자료", "I220"),
    ]
    checks: list[ExcelParityCheck] = []
    for name, python_value, sheet, cell in cells:
        excel_value = workbook[sheet][cell].value
        if excel_value is None:
            raise ValueError(f"{sheet}!{cell} 값이 비어 있습니다.")
        checks.append(ExcelParityCheck(name, float(python_value), float(excel_value)))
    return checks


def assert_default_parity(result: dict, tolerance: float = 1e-6) -> None:
    failures = [check for check in default_parity_checks(result) if abs(check.diff) > tolerance]
    if failures:
        details = ", ".join(
            f"{item.name}: python={item.python_value}, excel={item.excel_value}, diff={item.diff}"
            for item in failures
        )
        raise AssertionError(f"치킨 기본값 검증 실패: {details}")
