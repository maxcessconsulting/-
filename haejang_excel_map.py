from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pure_model import AGE_COLUMNS, ModelInput, TIME_LABELS, calculate, default_input


WORKBOOK_PATH = Path("/workspace/user_files/01-pgm.xlsx")
INPUT_SHEET = "조사치 입력 장표"
OUTPUT_SHEET = "표지"
OUTPUT_DAILY_SALES_CELL = "D9"

REGION_TO_EXCEL_CODE = {
    "서울 경기": 1,
    "충청권": 2,
    "호남권": 3,
    "대구경북": 4,
    "부산 경남": 5,
    "강원권": 6,
}

ADMIN_UNIT_TO_EXCEL_CODE = {
    "시 단위": 1,
    "군 단위": 2,
}

OPERATION_TO_EXCEL_FLAGS = {
    "직영점": {"D117": 1, "E117": 0, "F117": 0},
    "가맹점": {"D117": 0, "E117": 1, "F117": 0},
    "위탁운영": {"D117": 0, "E117": 0, "F117": 1},
}

TRAFFIC_START_ROW = 65
TRAFFIC_START_COLUMN = "E"
DIRECT_COMPETITOR_START_ROW = 86
INDIRECT_COMPETITOR_START_ROW = 97
COMPETITOR_COLUMNS = {
    "name": "E",
    "area": "F",
    "distance": "G",
    "location": "H",
    "visibility": "I",
    "accessibility": "J",
    "floor": "K",
    "sides": "L",
    "frontage": "M",
    "facility": "N",
    "parking": "O",
    "price": "P",
}


@dataclass(frozen=True)
class ExcelInputCell:
    field: str
    sheet: str
    cell: str
    label: str
    required: bool = True


@dataclass(frozen=True)
class ExcelParityCheck:
    name: str
    python_value: float
    excel_value: float

    @property
    def diff(self) -> float:
        return self.python_value - self.excel_value


BASE_INPUT_CELLS = [
    ExcelInputCell("survey_month", INPUT_SHEET, "E8", "조사월"),
    ExcelInputCell("survey_day", INPUT_SHEET, "H8", "조사일", required=False),
    ExcelInputCell("weekday", INPUT_SHEET, "L8", "조사요일"),
    ExcelInputCell("is_24h", INPUT_SHEET, "P8", "24시간 영업 여부", required=False),
    ExcelInputCell("region", INPUT_SHEET, "G13", "지역 권역"),
    ExcelInputCell("admin_unit", INPUT_SHEET, "G14", "행정 단위"),
    ExcelInputCell("apartment_households", INPUT_SHEET, "E25", "아파트 세대수"),
    ExcelInputCell("total_households", INPUT_SHEET, "I25", "주택계"),
    ExcelInputCell("resident_population", INPUT_SHEET, "J25", "주거인구"),
    ExcelInputCell("worker_population", INPUT_SHEET, "L25", "직장인구"),
    ExcelInputCell("annual_income", INPUT_SHEET, "P25", "가구당 연간 소득"),
    ExcelInputCell("deposit", INPUT_SHEET, "D113", "임차보증금"),
    ExcelInputCell("goodwill", INPUT_SHEET, "E113", "영업권"),
    ExcelInputCell("monthly_rent", INPUT_SHEET, "F113", "월임대료"),
    ExcelInputCell("management_fee", INPUT_SHEET, "G113", "관리비"),
]


def base_input_cell_map() -> dict[str, str]:
    return {item.field: f"{item.sheet}!{item.cell}" for item in BASE_INPUT_CELLS}


def model_input_to_excel_updates(data: ModelInput) -> dict[str, Any]:
    updates: dict[str, Any] = {
        "E8": data.survey_month,
        "L8": data.weekday,
        "G13": REGION_TO_EXCEL_CODE.get(data.region),
        "G14": ADMIN_UNIT_TO_EXCEL_CODE.get(data.admin_unit),
        "E25": data.apartment_households,
        "I25": data.total_households,
        "J25": data.resident_population,
        "L25": data.worker_population,
        "P25": data.annual_income,
        "D113": data.deposit,
        "E113": data.goodwill,
        "F113": data.monthly_rent,
        "G113": data.management_fee,
    }
    updates.update(OPERATION_TO_EXCEL_FLAGS.get(data.operation_type, OPERATION_TO_EXCEL_FLAGS["가맹점"]))

    for row_index, traffic_row in enumerate(data.traffic[: len(TIME_LABELS)], start=TRAFFIC_START_ROW):
        for col_index, value in enumerate(traffic_row[: len(AGE_COLUMNS)], start=5):
            updates[f"{chr(64 + col_index)}{row_index}"] = value

    updates.update(_competitor_updates(data.direct_competitors, DIRECT_COMPETITOR_START_ROW))
    updates.update(_competitor_updates(data.indirect_competitors, INDIRECT_COMPETITOR_START_ROW))
    return updates


def _competitor_updates(competitors, start_row: int) -> dict[str, Any]:
    updates: dict[str, Any] = {}
    for offset, competitor in enumerate(competitors[:10]):
        row = start_row + offset
        for attr, column in COMPETITOR_COLUMNS.items():
            updates[f"{column}{row}"] = getattr(competitor, attr)
    return updates


def read_stored_daily_sales(workbook_path: Path = WORKBOOK_PATH) -> float:
    workbook = load_workbook(workbook_path, data_only=True)
    value = workbook[OUTPUT_SHEET][OUTPUT_DAILY_SALES_CELL].value
    if value is None:
        raise ValueError(f"{OUTPUT_SHEET}!{OUTPUT_DAILY_SALES_CELL} 값이 비어 있습니다.")
    return float(value)


def default_parity_checks(workbook_path: Path = WORKBOOK_PATH) -> list[ExcelParityCheck]:
    workbook = load_workbook(workbook_path, data_only=True)
    result = calculate(default_input())
    cells = [
        ("daily_sales_thousand", result["daily_sales_thousand"], "표지", "D9"),
        ("daily_traffic", result["daily_traffic"], "기초 데이터(입력불가)", "C17"),
        ("main_customer_ratio", result["main_customer_ratio"], "기초 데이터(입력불가)", "E25"),
        ("twenties_ratio", result["twenties_ratio"], "기초 데이터(입력불가)", "L25"),
        ("traffic_potential_total", result["traffic_potential_total"], "SIMULATION(입력불가)", "B40"),
        ("household_potential_total", result["household_potential_total"], "SIMULATION(입력불가)", "D40"),
        ("worker_potential_total", result["worker_potential_total"], "SIMULATION(입력불가)", "E40"),
        ("candidate_score", result["candidate_score"], "기초 데이터(입력불가)", "G31"),
        ("total_competition_score", result["total_competition_score"], "SIMULATION(입력불가)", "E102"),
        ("month_index", result["month_index"], "기초자료", "N222"),
        ("weekday_index", result["weekday_index"], "기초자료", "I229"),
    ]
    checks: list[ExcelParityCheck] = []
    for name, python_value, sheet, cell in cells:
        excel_value = workbook[sheet][cell].value
        if excel_value is None:
            raise ValueError(f"{sheet}!{cell} 값이 비어 있습니다.")
        checks.append(ExcelParityCheck(name, float(python_value), float(excel_value)))
    return checks


def assert_default_parity(tolerance: float = 1e-3) -> None:
    failures = [check for check in default_parity_checks() if abs(check.diff) > tolerance]
    if failures:
        details = ", ".join(
            f"{item.name}: python={item.python_value}, excel={item.excel_value}, diff={item.diff}"
            for item in failures
        )
        raise AssertionError(f"해장국 기본값 검증 실패: {details}")
